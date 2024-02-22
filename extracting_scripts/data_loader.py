import os, sys
import json
from pathlib import Path as PATH
from classes import *
from openpyxl import load_workbook
from mapping import GEO_CODE, UNION_CODE, PAURASAVA_CODE, UPAZILLA_CODE, DISTRICT_CODE, DIVISION_CODE, NAME

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, 'bbs_docs')
OUTPUT_DIR = os.path.join(BASE_DIR, 'geocode_data')
p = PATH(BASE_DIR)

divisions = []
districts = []
upazillas = []
unions = []

def prepare_data(file_path):
    workbook = load_workbook(filename=file_path)
    sheets = workbook.sheetnames
    if len(sheets) > 1:
        for sheet in sheets:
            json_output= load_data(workbook[sheet])
            print(type(json_output ))
            # print(json_output)
            if json_output:
                count = 0
                # print(f"Division: {json_output['divisions']}")
                # print(f"District: {json_output['districts']}")
                # print(f"Upazilla: {json_output['upazillas']}")
                # print(f"Union: {json_output['unions']}")
                for union in json_output['unions']:
                    count += 1
                    print(count)
                # for parts in json_output:
                #     if (json_output[parts]):
                #         save_json(json_output[parts], parts)
                #     if (len(parts) > 0):
                #         save_json(json_output[parts], parts)
    else:
        json_output = load_data(workbook[sheets[0]])
        if json_output:
            count = 0
            for union in json_output['unions']:
                # if type(union['name']) is int:
                #     pass
                # if union['name'].startswith('WARD') an:
                #     pass
                if type(union['name']) is str and not union['name'].startswith('WARD'):
                    count += 1
                    # print(f'Union: {union}')
                    # print(f"Name: {union['name']}, Code: {union['geo_code']}")
            print(f'Total Unions {count}')
            # for parts in json_output:
            #     print(parts)
            #     save_json(json_output[parts], parts)
            # save_json(json_output, sheets[0])
        
def load_data(sheet):
    # Skipping the first row as it contains headers
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check for Division
        if row[0] and not any(row[1:4]) and row[5]:
            division_obj = {
                'type': 'Division',
                'geo_code': row[0],
                'name': row[5]
            }
            divisions.append(division_obj)
        # Check for District
        elif row[0] and row[1] and not row[2] and not row[3] and not row[4] and row[5]:
            district_obj = {
                'type': 'District',
                'geo_code': row[1],
                'division_geo_code': row[0],
                'name': row[5]
            }
            districts.append(district_obj)
        # Check for Upazilla
        elif row[0] and row[1] and row[2] and not row[3] and not row[4] and row[5]:
            upazilla_obj = {
                'type': 'Upazilla',
                'geo_code': row[2],
                'district_geo_code': row[1],
                'division_geo_code': row[0],
                'name': row[5]
            }
            upazillas.append(upazilla_obj)
        elif all(row[:5]) and row[5]:
            if type(row[5]) is str and not row[5].startswith('WARD'):
                union_obj={
                    # 'type': 'Union',
                    'geo_code': row[4],
                    # 'upazilla_geo_code': row[2],
                    # 'district_geo_code': row[1],
                    # 'division_geo_code': row[0],
                    'name': row[5] 
                }
                unions.append(union_obj)
        # Check
    return {
        'divisions': divisions,
        'districts': districts,
        'upazillas': upazillas,
        'unions': unions
    }



def save_json(data, sheet_name):
    output_file_path = os.path.join(OUTPUT_DIR, f'{sheet_name}.json')
    with open(output_file_path, 'w') as json_file:
        json.dump(data, json_file, indent=2)

# print(f'It is running')
for x in p.iterdir():
    if x.is_dir() and x.name == 'bbs_docs':
        for y in x.iterdir():
            if y.is_file() and y.name=='ctg.xlsx':
                prepare_data(y)